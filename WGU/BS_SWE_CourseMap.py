import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Define the course data
courses = [
    ["D278", "Scripting and Programming - Foundations", 3, "Yes", "Yes", "Yes", "Yes", "Programming"],
    ["D335", "Introduction to Programming in Python", 3, "Yes", "Yes", "Yes", "Programming"],
    ["D370", "IT Leadership Foundations", 3, "Yes", "Yes", "Yes", "Programming"],
    ["D276", "Web Development Foundations", 3, "Yes", "Yes", "Yes", "Programming"],
    ["D197", "Version Control", 1, "Yes", "Yes", "Yes", "Programming"],
    ["D277", "Front-End Web Development", 3, "Yes", "Yes", "Yes", "Programming"],
    ["C955", "Applied Probability and Statistics", 3, "Yes", "Yes", "Yes", "Programming"],
    ["D322", "Introduction to IT", 3, "Yes", "Yes", "Yes", "Yes", "General Education"],
    ["C683", "Natural Science Lab", 2, "Yes", "Yes", "Yes", "General Education"],
    ["C957", "Applied Algebra", 3, "Yes", "Yes", "Yes", "General Education"],
    ["C949", "Data Structures and Algorithms I", 4, "Yes", "Yes", "Yes", "General Education"],
    ["D372", "Introduction to Systems Thinking", 3, "Yes", "Yes", "Yes", "General Education"],
    ["D426", "Data Management - Foundations", 3, "Yes", "Yes", "Yes", "General Education"],
    ["D427", "Data Management - Applications", 4, "Yes", "Yes", "Yes", "General Education"],
    ["D280", "JavaScript Programming", 3, "Yes", "Yes", "Yes", "General Education"],
    ["D333", "Ethics in Technology", 3, "Yes", "Yes", "Yes", "General Education"],
    ["D326", "Advanced Data Management", 3, "Yes", "Yes", "Yes", "General Education"],
    ["D279", "User Interface Design", 3, "Yes", "Yes", "Yes", "General Education"],
    ["D479", "User Experience Design", 3, "Yes", "Yes", "Yes", "General Education"],
    ["D286", "Java Fundamentals", 3, "Yes", "Yes", "Yes", "General Education"],
    ["C458", "Health, Fitness and Wellness", 3, "Yes", "Yes", "Yes", "General Education"],
    ["D287", "Java Frameworks", 3, "Yes", "Yes", "Yes", "General Education"],
    ["D282", "Cloud Foundations", 3, "Yes", "Yes", "Yes", "General Education"],
    ["D386", "Hardware and Operating Systems Essentials", 3, "Yes", "Yes", "Yes", "General Education"],
    ["D324", "Business of IT - Project Management", 4, "Yes", "Yes", "Yes", "General Education"],
    ["D288", "Back-End Programming", 3, "Yes", "Yes", "Yes", "General Education"],
    ["D270", "Composition: Successful Self-Expression", 3, "Yes", "Yes", "Yes", "General Education"],
    ["D336", "Business of IT - Applications", 4, "Yes", "Yes", "Yes", "General Education"],
    ["D199", "Introduction to Physical and Human Geography", 3, "Yes", "Yes", "General Education"],
    ["D387", "Advanced Java", 3, "Yes", "Yes", "Yes", "General Education"],
    ["D385", "Software Security and Testing", 3, "Yes", "Yes", "General Education"],
    ["D339", "Technical Communication", 3, "Yes", "Yes", "Yes", "General Education"],
    ["D480", "Software Design and Quality Assurance", 3, "Yes", "Yes", "Yes", "General Education"],
    ["C963", "American Politics and the US Constitution", 3, "Yes", "Yes", "General Education"],
    ["D284", "Software Engineering", 4, "Yes", "Yes", "Yes", "General Education"],
    ["D308", "Mobile Application Development(Android)", 3, "Yes", "Yes", "General Education"],
    ["D424", "Software Engineering Capstone", 4, "Yes", "Yes", "General Education"]

]

# Create a new workbook and select the active sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BS Software Engineering"

# Add header row
header = ["ID", "Course Name", "Units", "Study.com Transfers", "Straighterline", "Sophia.org", "Certificate", "Category"]
ws.append(header)

# Apply styles to the header row
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")

for col in range(1, len(header) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# Add course data
for course in courses:
    ws.append(course)

# Adjust column widths for better readability
column_widths = [10, 40, 8, 15, 20, 20, 20, 15, 20]
for i, width in enumerate(column_widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

# Save the workbook
wb.save("BS_Software_Engineering_Course_Tracker.xlsx")

print("Spreadsheet created successfully!")
